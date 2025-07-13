import openpyxl
from openpyxl.styles import Font

class Proyecto:
    def __init__(self, id, nombre, descripcion):
        self.__id = id
        self.__nombre = nombre
        self.__descripcion = descripcion
        self.__colaboradores = []
        self.__actividades = []

    def asignar_colaborador(self, colaborador):
        if colaborador not in self.__colaboradores:
            self.__colaboradores.append(colaborador)
            print(f'colaborador {colaborador.get_nombre()} asignado al proyecto {self.__nombre}.')
        else:
            print(f'Colaborador ya asignado a este proyecto.')

    def agregar_actividad(self, actividad):
        if actividad is not None:
            self.__actividades.append(actividad)
        else:
            print("Actividad inválida.")

    def obtener_total_horas(self):
        return sum(act.calcular_duracion() for act in self.__actividades if act.esta_activa())

    def mostrar_resumen(self):
        print(f'Proyecto: {self.__nombre}')
        print(f'Descripción: {self.__descripcion}')
        print(f'Colaboradores asignados: {len(self.__colaboradores)}')
        print(f'Total de actividades registradas: {len(self.__actividades)}')
        print(f'Total de horas trabajadas: {self.obtener_total_horas():.2f} h')
        
    def exportar_resumen_excel(self, archivo=None):
        if archivo is None:
            archivo = f'resumen_proyecto_{self.__nombre}.xlsx'
    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen del Proyecto"
    
        #nombre
        ws['A1'] = f'Resumen del Proyecto: {self.__nombre}'
        ws['A1'].font = Font(bold=True, size=14)
    

        ws.append([])
        ws.append(["ID del Proyecto", self.__id])
        ws.append(["Nombre", self.__nombre])
        ws.append(["Descripción", self.__descripcion])
        ws.append(["Colaboradores asignados", len(self.__colaboradores)])
        ws.append(["Actividades registradas", len(self.__actividades)])
        ws.append(["Total de horas trabajadas", round(self.obtener_total_horas(), 2)])
    
        ws.append([])
        ws.append(["Fecha", "Descripción", "Tipo", "Colaborador", "Horas"])
    
        for act in self.__actividades:
            if act.esta_activa():
                ws.append([
                    act.get_fecha().isoformat(),
                    act.get_descripcion(),
                    act.get_tipo(),
                    act.get_usuario().get_nombre() + ' ' + act.get_usuario().get_apellido(),
                    round(act.calcular_duracion(), 2)
                ])
    
        #guardar archivo
        wb.save(archivo)
        print(f"Resumen exportado a Excel: {archivo}")

    # getters
    def get_id(self):
        return self.__id

    def get_nombre(self):
        return self.__nombre

    def get_descripcion(self):
        return self.__descripcion

    def get_colaboradores(self):
        return self.__colaboradores

    def get_actividades(self):
        return self.__actividades
